from fastapi import APIRouter, Request, UploadFile, File, Form, Depends, HTTPException
from fastapi.responses import HTMLResponse
import io
import openpyxl
import json
import datetime
import uuid
from config import TEMPLATES
from core.redis import redis_client
from core.logging import setup_logger
from auth import verificar_usuario


def parse_form_dict(form: dict[str, str], prefixo: str) -> dict[str, str]:
    """
    Extrai um dicionário de campos do FormData no formato: prefixo[chave] = valor

    Exemplo:
        extras[nome] => João
        extras[idade] => 25
        =>
        {"nome": "João", "idade": "25"}
    """
    resultado = {}

    for key, value in form.items():
        if key.startswith(f"{prefixo}[") and key.endswith("]"):       
            if len(value) > 200:
                raise HTTPException(status_code=400, detail=f"Valor: {value} da chave: {key} - maior que 200 caracteres")
            chave = key[len(prefixo) + 1:-1]  # Remove prefixo[ e ]
            resultado[chave] = value

    return resultado
 

router = APIRouter()
logger = setup_logger("qa", "routes")


@router.get("/qa", response_class=HTMLResponse, dependencies=[Depends(verificar_usuario)])
async def pagina_multi_testes(request: Request):
    context = {
        "request": request
    }
    return TEMPLATES.TemplateResponse("whats_multi.html", context=context)

@router.get("/qa/webhook", response_class=HTMLResponse, dependencies=[Depends(verificar_usuario)])
async def pagina_multi_testes_webhook(request: Request):
    context = {
        "request": request
    }
    return TEMPLATES.TemplateResponse("whats_webhook.html", context=context)

@router.get("/qa/manual", response_class=HTMLResponse)
async def pagina_testes_manuais(request: Request, usuario = Depends(verificar_usuario)):
    context = {
        "request": request,
        "numeros_disponiveis": usuario["numeros_disponiveis"]
    }
    return TEMPLATES.TemplateResponse("whats_manual.html", context=context)

@router.get("/qa/historico", response_class=HTMLResponse)
async def pagina_historico_testes(request: Request, usuario = Depends(verificar_usuario)):
    dados_teste = redis_client.get(f"canal:{usuario.id}")
    if not dados_teste:
        return

    dados_teste = json.loads(dados_teste)
    passos = dados_teste["passos"]
    testes_jsons = redis_client.lrange("historico_testes", 0, -1)
    testes = [json.loads(t) for t in testes_jsons]

    return TEMPLATES.TemplateResponse("historico_testes.html", {
        "request": request,
        "testes": testes
    })

@router.post("/webhook")
async def receber_webhook(payload: dict):
    """Recebe um Webhook e publica no canal correto do Redis"""
    mensagem_obj = {
        "status": "pendente"
    }
    arquivo_id = payload.get("conversationId")
    mensagem_obj["arquivo"] = arquivo_id
    logger.info(f"📥 Webhook recebido. Canal: {arquivo_id}: {payload}")
    tipo_mensagem:str = payload["output"][0].get("response_type")
    match tipo_mensagem:
        case 'text':
            mensagem_recebida: str = payload["output"][0].get("text")
        case 'image':
            mensagem_recebida: str = payload["output"][0].get("source")
        case 'option':
            mensagem_recebida: str = payload["output"][0].get("title")
            options: list = payload["output"][0]["options"]
            botoes = [{objeto_botao["label"]: objeto_botao["value"]["input"]["text"] for objeto_botao in options}]
            mensagem_obj["options"] = botoes
        case _:
            pass

    mensagem_recebida = "\n".join(linha.strip() for linha in mensagem_recebida.strip().splitlines())
    mensagem_obj["mensagem"] = mensagem_recebida
    timestamp = datetime.datetime.today().strftime("%Y/%m/%d-%H:%M:%S")
    mensagem_obj["timestamp"] = timestamp

    # Notificar frontend via WebSocket
    redis_client.publish(f"canal:{arquivo_id}", json.dumps(mensagem_obj))

    return {"message": "Webhook processado com sucesso!", "resultado": 'pendente'}

@router.post("/qa/enviar-multi-teste")
async def enviar_teste(
    file: UploadFile = File(...),
    nome: str = Form(...),
    telefone: str = Form(...)
):
    """Recebe um arquivo Excel, obtém um arquivo_id do APP externo de cada planilha de teste e armazena os passos no Redis."""

    # 📌 1️⃣ Lendo o conteúdo do arquivo como um buffer
    file_content = await file.read()
    file_stream = io.BytesIO(file_content)  # Convertendo para stream de bytes

    # 📌 2️⃣ Carregando o arquivo diretamente no `openpyxl`
    workbook = openpyxl.load_workbook(file_stream)
    planilhas = {}

    for index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]
        passos = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Pulando cabeçalho
            id_passo, tipo, valor, validar = row
            if id_passo and tipo and valor:
                passos.append({
                    "id": id_passo,
                    "tipo": tipo.lower().strip(),  # "enviar" ou "receber" ou "esperar"
                    "valor": valor,
                    "validar": validar,
                    "status": "pendente"
                })

        # 🔹 1️⃣ Solicita o arquivo_id ao APP externo com a mensagem inicial
        #data = chamar_api_externa("Teste", nome, str(int(telefone)+index))
        arquivo_id = uuid.uuid4()
        planilhas[sheet_name] = arquivo_id
        # 🔹 3️⃣ Salva os passos no Redis
        redis_client.setex(
            f"canal:{arquivo_id}", 3600, 
            json.dumps({
                "status": "pendente",
                "name": sheet_name,
                "arquivo": file.filename,
                "passos": passos,
                "nome": nome,
                "telefone": str(int(telefone)+index)
            })
        )

    return {
        "planilhas": planilhas,
        "message": "Testes iniciados!",
        "total_testes": len(planilhas.keys())
    }

@router.post("/qa/criar-testes-multiplos-manuais")
async def criar_multiplos_testes_manuais(
    request: Request,
    usuario = Depends(verificar_usuario)):
    """
    Gera múltiplas instâncias de teste manualmente com variação de telefone.
    """
    form = await request.form()
    numeros = form.getlist("numeros")
    extras = parse_form_dict(form, prefixo="extras")

    invalidos = [n for n in numeros if n not in usuario["numeros_disponiveis"]]
    if invalidos:
        raise HTTPException(status_code=400, detail=f"Números inválidos: {invalidos}")

    resultados = {}

    for numero in numeros:
        arquivo_id = str(uuid.uuid4())[:8]      

        redis_client.setex(
            f"wss:{arquivo_id}", 3600,
            json.dumps({
                "nome": usuario["user"]["name"],
                "telefone": numero,
                "extras": extras
            })
        )

        resultados[numero] = arquivo_id
    
    return {"message": "Testes criados com sucesso", "testes": resultados}

