from fastapi import FastAPI, WebSocket, WebSocketDisconnect, File, UploadFile, Form, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
import asyncio
import json
import uuid
from dotenv import load_dotenv
import openpyxl
import redis
import io
import os
import requests
import datetime
import logging
import websockets


def chamar_api_externa(message: str, name: str, phone: str, conversation_id: str = None) -> dict:
    url = f"https://{MESSAGE_URL}"
    body = {
    "text": message,
    "context": {
        "name": name,
        "phone": phone
    }
    }

    if conversation_id:
        body['conversationId'] = conversation_id

    payload = json.dumps(body)
    headers = {
    'x-api-key': TOKEN,
    'Content-Type': 'application/json'
    }

    response = requests.post(url, headers=headers, data=payload)
    if response.ok:
        return response.json()
    else:
        raise Exception(f'Falha na API({response.status_code})')


load_dotenv()
EXTERNAL_WS_URL = os.environ.get("EXTERNAL_WS_URL")
TOKEN = os.environ.get("TOKEN")
MESSAGE_URL = os.environ.get("MESSAGE_URL")

# Configuração do Logger
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Configuração do APP FastAPI
app = FastAPI(title="Automação de Testes")
app.mount("/static", StaticFiles(directory="static"), name="static")
TEMPLATES = Jinja2Templates(directory="templates")

@app.middleware("http")
async def log_unmatched_requests(request: Request, call_next):
    response = await call_next(request)

    # Se o status for 404 ou 405, significa que a rota não foi encontrada ou o método não é permitido
    if response.status_code in [404, 405]:
        body = await request.body()
        
        logger.warning(f"🔍 Requisição não mapeada: {request.method} {request.url}")
        logger.warning(f"📥 Headers: {dict(request.headers)}")
        logger.warning(f"📦 Body: {body.decode() if body else None}")

        return JSONResponse(
            status_code=405,
            content={"detail": "Método não permitido ou rota não encontrada"},
        )

    return response

# Conexão com Redis
redis_client = redis.Redis(host="localhost", port=6379, decode_responses=True)

@app.get("/testes", response_class=HTMLResponse)
async def pagina_multi_testes(request: Request):
    context = {
        "request": request
    }
    return TEMPLATES.TemplateResponse("whats_multi.html", context=context)

@app.post("/webhook")
async def receber_webhook(payload: dict):
    """Recebe um Webhook e publica no canal correto do Redis"""
    arquivo_id = payload.get("conversationId")
    mensagem_recebida: str = payload["output"][0].get("text") if payload["output"][0].get("response_type") == "text" else payload["output"][0].get("title")
    mensagem_recebida = "\n".join(linha.strip() for linha in mensagem_recebida.strip().splitlines())
    timestamp = datetime.datetime.today().strftime("%Y/%m/%d-%H:%M:%S")

    # Buscar os passos do teste no Redis
    dados_teste = redis_client.get(f"canal:{arquivo_id}")

    if not dados_teste:
        return {"error": "Teste não encontrado"}  

    redis_client.setex(f"canal:{arquivo_id}", 3600, dados_teste)

    # Notificar frontend via WebSocket
    redis_client.publish(f"canal:{arquivo_id}", json.dumps({
        "arquivo": arquivo_id,
        "status": 'pendente',
        "timestamp": timestamp,
        "mensagem": mensagem_recebida
    }))

    return {"message": "Webhook processado com sucesso!", "resultado": 'pendente'}

@app.post("/enviar-multi-teste")
async def enviar_teste(
    file: UploadFile = File(...),
    nome: str = Form(...),
    telefone: str = Form(...)
):
    """Recebe um arquivo Excel, obtém um arquivo_id do APP externo de cada planilha de teste e armazena os passos no Redis."""

    # 📌 1️⃣ Lendo o conteúdo do arquivo como um buffer
    file_content = await file.read()
    file_stream = io.BytesIO(file_content)  # Convertendo para stream de bytes

    # 📌 2️⃣ Carregando o arquivo diretamente no `openpyxl`
    workbook = openpyxl.load_workbook(file_stream)
    planilhas = {}

    for index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]
        passos = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Pulando cabeçalho
            id_passo, tipo, valor, validar = row
            if id_passo and tipo and valor:
                passos.append({
                    "id": id_passo,
                    "tipo": tipo.lower().strip(),  # "enviar" ou "receber" ou "esperar"
                    "valor": valor,
                    "validar": validar,
                    "status": "pendente"
                })

        # 🔹 1️⃣ Solicita o arquivo_id ao APP externo com a mensagem inicial
        data = chamar_api_externa("Teste", nome, str(int(telefone)+index))
        arquivo_id = data.get("conversationId")
        planilhas[sheet_name] = arquivo_id
        # 🔹 3️⃣ Salva os passos no Redis
        redis_client.setex(
            f"canal:{arquivo_id}", 3600, 
            json.dumps({
                "status": "pendente",
                "name": sheet_name,
                "arquivo": file.filename,
                "passos": passos,
                "nome": nome,
                "telefone": telefone
            })
        )

    return {
        "planilhas": planilhas,
        "message": "Testes iniciados!",
        "total_testes": len(planilhas.keys())
    }

@app.websocket("/ws/notificacoes/{arquivo_id}")
async def websocket_notificacoes(websocket: WebSocket, arquivo_id: str):
    """WebSocket escuta mensagens do Redis e processa os passos um por um, garantindo a ordem correta"""
    await websocket.accept()
    
    pubsub = redis_client.pubsub()
    canal = f"canal:{arquivo_id}"
    pubsub.subscribe(canal)

    # 🔹 Buscamos os passos no Redis
    dados_teste = redis_client.get(f"canal:{arquivo_id}")
    if not dados_teste:
        return

    dados_teste = json.loads(dados_teste)
    passos = dados_teste["passos"]
    
    try:
        for passo in passos:  # 🔹 Agora processamos os passos um por um
            if passo["tipo"] == "enviar" and passo["status"] == "pendente":
                logger.info(f"📤 Enviando mensagem para API externa: {passo['valor']}")
                chamar_api_externa(passo["valor"], dados_teste["nome"], dados_teste["telefone"], arquivo_id)
                passo["status"] = "enviado"

                # 🔹 Atualiza o Redis com o novo status do passo
                redis_client.setex(f"canal:{arquivo_id}", 3600, json.dumps(dados_teste))

                # 🔹 Envia diretamente ao WebSocket (além de publicar no Redis)
                mensagem_ws = {
                    "arquivo": arquivo_id,
                    "status": "enviado",
                    "timestamp": datetime.datetime.today().strftime("%Y/%m/%d-%H:%M:%S"),
                    "mensagem": str(passo["valor"])
                }
                await websocket.send_text(json.dumps(mensagem_ws))

            elif passo["tipo"] == "receber" and passo["status"] == "pendente":
                while True:  # 🔹 Aguarda a resposta correta antes de avançar
                    mensagem = pubsub.get_message(ignore_subscribe_messages=True, timeout=5)
                    if not mensagem:
                        continue

                    data = json.loads(mensagem["data"])
                    mensagem_recebida = data['mensagem']
                    mensagem_esperada:str = passo["valor"]
                    mensagem_esperada = "\n".join(linha.strip() for linha in mensagem_esperada.strip().splitlines())
                    
                    if passo["validar"] == "exato" and mensagem_recebida == mensagem_esperada:
                        resultado = "success"
                    elif passo["validar"] == "contém" and mensagem_esperada in mensagem_recebida:
                        resultado = "success"
                    else:
                        resultado = "error"
                        logger.warning(f"📩 Mensagem recebida: |{mensagem_recebida}|\nMensagem esperada: |{mensagem_esperada}|")

                    # 🔹 Atualiza o status do passo no Redis
                    passo["status"] = resultado
                    redis_client.setex(f"canal:{arquivo_id}", 3600, json.dumps(dados_teste))

                    # 🔹 Envia diretamente ao WebSocket (além de publicar no Redis)
                    mensagem_ws = {
                        "arquivo": arquivo_id,
                        "status": resultado,
                        "timestamp": datetime.datetime.today().strftime("%Y/%m/%d-%H:%M:%S"),
                        "mensagem": mensagem_recebida
                    }
                    await websocket.send_text(json.dumps(mensagem_ws))

                    break  # 🔹 Só avança para o próximo passo quando esse estiver resolvido
            
            elif passo["tipo"] == "esperar" and passo["status"] == "pendente":
                logger.info(f"💤 Robô dormindo: {passo['valor']} segundos")
                await asyncio.sleep(int(passo["valor"]))

                # 🔹 Atualiza o status do passo no Redis
                passo["status"] = resultado
                redis_client.setex(f"canal:{arquivo_id}", 3600, json.dumps(dados_teste))
                
        mensagem_ws = {
            "arquivo": arquivo_id,
            "status": "end",
            "timestamp": datetime.datetime.today().strftime("%Y/%m/%d-%H:%M:%S")
        }
        await websocket.send_text(json.dumps(mensagem_ws))
    except WebSocketDisconnect:
        logger.error(f"🔴 WebSocket {arquivo_id} desconectado. Limpando conexões...")

        # 🔹 Remove assinaturas do Redis
        pubsub.unsubscribe(canal)

        # 🔹 Fecha conexões pendentes
        redis_client.delete(f"canal:{arquivo_id}")

        # 🔹 Log de finalização
        logger.error(f"✅ Conexões para {arquivo_id} foram encerradas.")

    except Exception as e:
        logger.error(f"⚠️ Erro no WebSocket: {str(e)}")
    finally:
        await websocket.close()

@app.websocket("/ws/notificacao/{arquivo_id}")
async def websocket_notificacao(websocket: WebSocket, arquivo_id: str):
    """WebSocket que processa o arquivo Excel e realiza os testes automaticamente."""
    await websocket.accept()

    logger.info("✅ Cliente conectado ao WebSocket de Testes")
    CONVERSATION_ID = str(uuid.uuid1())
    external_ws_url = f"wss://{EXTERNAL_WS_URL}?conversationId={CONVERSATION_ID}&token={TOKEN}"

    try:
        pubsub = redis_client.pubsub()
        canal = f"canal:{arquivo_id}"
        pubsub.subscribe(canal)

        # 🔹 Buscamos os passos no Redis
        dados_teste = redis_client.get(f"canal:{arquivo_id}")
        if not dados_teste:
            return

        dados_cliente = json.loads(dados_teste)
        passos = dados_cliente.get("passos")
        nome = dados_cliente.get("nome")
        telefone = dados_cliente.get("telefone")

        async with websockets.connect(external_ws_url) as external_ws:
            mensagem_envio = {
                        "action": "userToKloe",
                        "conversationId": CONVERSATION_ID,
                        "data": {
                            "message": "Teste",
                            "type": "text",
                            "user": {
                                "name": nome,
                                "phone": telefone,
                                "id": CONVERSATION_ID,
                                "ra": CONVERSATION_ID
                            },
                            "bot": "Kloe",
                            "system": False,
                            "token": TOKEN
                        }
                    }
            await external_ws.send(json.dumps(mensagem_envio))
            logger.info(f"📤 Mensagem enviada: Teste")
            for passo in passos:
                if passo["tipo"] == "enviar":
                    mensagem_recebida = str(passo['valor'])
                    mensagem_envio = {
                        "action": "userToKloe",
                        "conversationId": CONVERSATION_ID,
                        "data": {
                            "message": mensagem_recebida,
                            "type": "text",
                            "user": {
                                "name": nome,
                                "phone": telefone,
                                "id": CONVERSATION_ID,
                                "ra": CONVERSATION_ID
                            },
                            "bot": "Kloe",
                            "system": False,
                            "token": TOKEN
                        }
                    }
                    await external_ws.send(json.dumps(mensagem_envio))
                    logger.info(f"📤 Mensagem enviada: {mensagem_recebida}")
                    mensagem_ws = {
                        "arquivo": arquivo_id,
                        "status": "processando",
                        "timestamp": datetime.datetime.today().strftime("%Y/%m/%d-%H:%M:%S"),
                        "mensagem": mensagem_recebida
                    }
                    await websocket.send_text(json.dumps(mensagem_ws))
                elif passo["tipo"] == "receber":
                    try:
                        while True:
                            resposta = await external_ws.recv()
                            resposta_json = json.loads(resposta)
                            if resposta_json['action'] == 'kloeToUser':
                                break
                        mensagem_recebida:str = resposta_json["data"]["messages"][0]["text"]
                        mensagem_esperada:str = str(passo["valor"])
                        mensagem_recebida = "\n".join(linha.strip() for linha in mensagem_recebida.strip().splitlines())
                        mensagem_esperada = "\n".join(linha.strip() for linha in mensagem_esperada.strip().splitlines())
                        if passo["validar"] == "exato" and mensagem_recebida == mensagem_esperada:
                            resultado = "success"
                        elif passo["validar"] == "contém" and mensagem_esperada in mensagem_recebida:
                            resultado = "success"
                        else:
                            resultado = "error"
                            logger.warning(f"📩 Mensagem recebida: |{mensagem_recebida}|\nMensagem esperada: |{mensagem_esperada}|")

                        mensagem_ws = {
                            "arquivo": arquivo_id,
                            "status": resultado,
                            "timestamp": datetime.datetime.today().strftime("%Y/%m/%d-%H:%M:%S"),
                            "mensagem": mensagem_recebida
                        }
                        await websocket.send_text(json.dumps(mensagem_ws))
                        logger.info(f"📩 Mensagem recebida: {True} | Resultado: {resultado}")

                    except asyncio.TimeoutError:
                        mensagem_ws = {
                            "arquivo": arquivo_id,
                            "status": "error",
                            "timestamp": datetime.datetime.today().strftime("%Y/%m/%d-%H:%M:%S"),
                            "mensagem": "Timeout na espera da resposta."
                        }
                        await websocket.send_text(json.dumps(mensagem_ws))
                        logger.error("❌ Timeout na espera da resposta.")
                elif passo["tipo"] == "esperar":
                    await asyncio.sleep(int(passo["valor"]))
                    logger.info(f"⏳ Esperou {passo['valor']} segundos")

            logger.info("✅ Teste finalizado!")
            mensagem_ws = {
            "arquivo": arquivo_id,
            "status": "end",
            "timestamp": datetime.datetime.today().strftime("%Y/%m/%d-%H:%M:%S")
            }
            await websocket.send_text(json.dumps(mensagem_ws))
    except WebSocketDisconnect:
        logger.error(f"🔴 WebSocket {arquivo_id} desconectado. Limpando conexões...")

        # 🔹 Remove assinaturas do Redis
        pubsub.unsubscribe(canal)

        # 🔹 Fecha conexões pendentes
        redis_client.delete(f"canal:{arquivo_id}")

        # 🔹 Log de finalização
        logger.error(f"✅ Conexões para {arquivo_id} foram encerradas.")

    except Exception as e:
        logger.error(f"⚠️ Erro no WebSocket: {str(e)}")
    finally:
        await websocket.close()
